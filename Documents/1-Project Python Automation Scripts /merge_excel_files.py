#!/usr/bin/env python3
"""
Excel Files Merger
Merges all Excel files in a specified folder into a single master Excel file.
"""

import os
import sys
from datetime import datetime
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
import logging

# Set up logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


class ExcelMerger:
    """Class to handle merging of multiple Excel files."""
    
    def __init__(self, input_folder, output_folder=None):
        """
        Initialize the Excel Merger.
        
        Args:
            input_folder (str): Path to folder containing Excel files
            output_folder (str): Path to output folder (defaults to input_folder)
        """
        self.input_folder = Path(input_folder)
        self.output_folder = Path(output_folder) if output_folder else self.input_folder
        self.merged_data = []
        
        # Validate input folder
        if not self.input_folder.exists():
            raise ValueError(f"Input folder does not exist: {self.input_folder}")
        
        # Create output folder if it doesn't exist
        self.output_folder.mkdir(parents=True, exist_ok=True)
    
    def get_excel_files(self):
        """
        Get all Excel files from the input folder.
        
        Returns:
            list: List of Excel file paths
        """
        excel_extensions = ['.xlsx', '.xls', '.xlsm']
        excel_files = []
        
        for ext in excel_extensions:
            excel_files.extend(self.input_folder.glob(f'*{ext}'))
        
        logger.info(f"Found {len(excel_files)} Excel file(s) in {self.input_folder}")
        return sorted(excel_files)
    
    def read_excel_file(self, file_path):
        """
        Read all sheets from an Excel file.
        
        Args:
            file_path (Path): Path to the Excel file
            
        Returns:
            dict: Dictionary of DataFrames (sheet_name: DataFrame)
        """
        try:
            # Read all sheets
            excel_file = pd.ExcelFile(file_path)
            sheets_data = {}
            
            for sheet_name in excel_file.sheet_names:
                try:
                    df = pd.read_excel(file_path, sheet_name=sheet_name)
                    if not df.empty:
                        sheets_data[sheet_name] = df
                        logger.info(f"  - Read sheet '{sheet_name}': {len(df)} rows")
                    else:
                        logger.warning(f"  - Sheet '{sheet_name}' is empty, skipping")
                except Exception as e:
                    logger.error(f"  - Error reading sheet '{sheet_name}': {str(e)}")
                    continue
            
            return sheets_data
            
        except Exception as e:
            logger.error(f"Error reading file {file_path.name}: {str(e)}")
            return {}
    
    def process_files(self):
        """
        Process all Excel files and merge them.
        
        Returns:
            pd.DataFrame: Merged DataFrame
        """
        excel_files = self.get_excel_files()
        
        if not excel_files:
            logger.warning("No Excel files found to merge!")
            return pd.DataFrame()
        
        all_dataframes = []
        
        for file_path in excel_files:
            logger.info(f"Processing: {file_path.name}")
            
            sheets_data = self.read_excel_file(file_path)
            
            for sheet_name, df in sheets_data.items():
                # Add source file information
                df['Source_File'] = file_path.name
                df['Source_Sheet'] = sheet_name
                df['Processed_Date'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                
                all_dataframes.append(df)
        
        if all_dataframes:
            merged_df = pd.concat(all_dataframes, ignore_index=True, sort=False)
            logger.info(f"\nTotal merged records: {len(merged_df)}")
            logger.info(f"Total columns: {len(merged_df.columns)}")
            return merged_df
        else:
            return pd.DataFrame()
    
    def save_merged_file(self, merged_df, add_formatting=True):
        """
        Save the merged DataFrame to an Excel file with optional formatting.
        
        Args:
            merged_df (pd.DataFrame): Merged DataFrame
            add_formatting (bool): Whether to add formatting to the output file
            
        Returns:
            str: Path to the saved file
        """
        if merged_df.empty:
            logger.warning("No data to save!")
            return None
        
        # Generate output filename with timestamp
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        output_filename = f"merged_excel_{timestamp}.xlsx"
        output_path = self.output_folder / output_filename
        
        # Save to Excel
        logger.info(f"\nSaving merged file to: {output_path}")
        
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            merged_df.to_excel(writer, sheet_name='Merged_Data', index=False)
            
            if add_formatting:
                # Access the workbook and worksheet
                workbook = writer.book
                worksheet = writer.sheets['Merged_Data']
                
                # Format header row
                header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                header_font = Font(bold=True, color='FFFFFF', size=11)
                
                for cell in worksheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Auto-adjust column widths
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    
                    for cell in column:
                        try:
                            if cell.value:
                                max_length = max(max_length, len(str(cell.value)))
                        except:
                            pass
                    
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
                
                # Freeze the header row
                worksheet.freeze_panes = 'A2'
        
        logger.info(f"Successfully saved merged file!")
        return str(output_path)
    
    def merge(self, add_formatting=True):
        """
        Main method to merge Excel files.
        
        Args:
            add_formatting (bool): Whether to add formatting to output
            
        Returns:
            str: Path to merged file or None if failed
        """
        logger.info("=" * 60)
        logger.info("Excel Files Merger - Starting Process")
        logger.info("=" * 60)
        
        merged_df = self.process_files()
        
        if merged_df.empty:
            logger.error("No data was merged. Exiting.")
            return None
        
        output_path = self.save_merged_file(merged_df, add_formatting)
        
        logger.info("=" * 60)
        logger.info("Merge Complete!")
        logger.info("=" * 60)
        
        return output_path


def main():
    """Main execution function."""
    # Default configuration
    INPUT_FOLDER = "./excel_files"
    OUTPUT_FOLDER = "./output"
    
    # You can also accept command-line arguments
    if len(sys.argv) > 1:
        INPUT_FOLDER = sys.argv[1]
    if len(sys.argv) > 2:
        OUTPUT_FOLDER = sys.argv[2]
    
    try:
        merger = ExcelMerger(INPUT_FOLDER, OUTPUT_FOLDER)
        output_file = merger.merge(add_formatting=True)
        
        if output_file:
            print(f"\n✓ Success! Merged file created at: {output_file}")
        else:
            print("\n✗ Failed to create merged file.")
            sys.exit(1)
            
    except Exception as e:
        logger.error(f"Fatal error: {str(e)}")
        print(f"\n✗ Error: {str(e)}")
        sys.exit(1)


if __name__ == "__main__":
    main()
