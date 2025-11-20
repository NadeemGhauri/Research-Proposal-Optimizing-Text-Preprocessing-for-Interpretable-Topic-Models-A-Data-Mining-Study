#!/usr/bin/env python3
"""
Excel Report Generator
Automated report generation with pivot tables, charts, and email delivery.
"""

import os
import sys
import json
import smtplib
from pathlib import Path
from datetime import datetime
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders
import logging

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo

# Set up logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


class ExcelReportGenerator:
    """Class to handle automated Excel report generation."""
    
    def __init__(self, config_file=None):
        """
        Initialize the Excel Report Generator.
        
        Args:
            config_file (str): Path to JSON configuration file
        """
        self.config = self._load_config(config_file)
        self.data_sources = {}
        self.workbook = None
        self.output_file = None
        
        # Create output directory
        output_dir = Path(self.config.get('output_folder', './reports'))
        output_dir.mkdir(parents=True, exist_ok=True)
        self.output_folder = output_dir
    
    def _load_config(self, config_file):
        """Load configuration from JSON file or use defaults."""
        default_config = {
            'output_folder': './reports',
            'report_name': 'automated_report',
            'data_sources': [],
            'kpis': [],
            'charts': [],
            'conditional_formatting': [],
            'password': None,
            'email': {
                'enabled': False,
                'smtp_server': 'smtp.gmail.com',
                'smtp_port': 587,
                'sender_email': '',
                'sender_password': '',
                'recipients': [],
                'subject': 'Automated Excel Report',
                'body': 'Please find the attached report.'
            }
        }
        
        if config_file and Path(config_file).exists():
            try:
                with open(config_file, 'r') as f:
                    user_config = json.load(f)
                default_config.update(user_config)
                logger.info(f"Loaded configuration from: {config_file}")
            except Exception as e:
                logger.warning(f"Error loading config file: {e}. Using defaults.")
        
        return default_config
    
    def load_data_source(self, source_config):
        """
        Load data from a source (CSV, Excel, JSON).
        
        Args:
            source_config (dict): Source configuration with 'name', 'type', and 'path'
        
        Returns:
            pd.DataFrame: Loaded data
        """
        name = source_config.get('name', 'data')
        source_type = source_config.get('type', 'csv')
        path = source_config.get('path')
        
        if not path or not Path(path).exists():
            logger.error(f"Data source not found: {path}")
            return pd.DataFrame()
        
        try:
            if source_type == 'csv':
                df = pd.read_csv(path)
            elif source_type == 'excel':
                sheet_name = source_config.get('sheet_name', 0)
                df = pd.read_excel(path, sheet_name=sheet_name)
            elif source_type == 'json':
                df = pd.read_json(path)
            else:
                logger.error(f"Unsupported source type: {source_type}")
                return pd.DataFrame()
            
            logger.info(f"Loaded '{name}': {len(df)} rows, {len(df.columns)} columns")
            return df
        
        except Exception as e:
            logger.error(f"Error loading data source '{name}': {e}")
            return pd.DataFrame()
    
    def load_all_data_sources(self):
        """Load all configured data sources."""
        logger.info("Loading data sources...")
        
        for source in self.config.get('data_sources', []):
            name = source.get('name', f'source_{len(self.data_sources)}')
            df = self.load_data_source(source)
            if not df.empty:
                self.data_sources[name] = df
        
        logger.info(f"Loaded {len(self.data_sources)} data source(s)")
    
    def create_data_sheet(self, sheet_name, dataframe, format_as_table=True):
        """
        Create a worksheet with data.
        
        Args:
            sheet_name (str): Name of the sheet
            dataframe (pd.DataFrame): Data to add
            format_as_table (bool): Format as Excel table
        """
        if dataframe.empty:
            logger.warning(f"No data to create sheet '{sheet_name}'")
            return None
        
        # Create or get worksheet
        if sheet_name in self.workbook.sheetnames:
            ws = self.workbook[sheet_name]
            self.workbook.remove(ws)
        
        ws = self.workbook.create_sheet(sheet_name)
        
        # Write data
        for r in dataframe_to_rows(dataframe, index=False, header=True):
            ws.append(r)
        
        # Format header
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=11)
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Format as table
        if format_as_table:
            tab = Table(displayName=sheet_name.replace(' ', '_'), ref=ws.dimensions)
            style = TableStyleInfo(
                name="TableStyleMedium9", 
                showFirstColumn=False,
                showLastColumn=False, 
                showRowStripes=True, 
                showColumnStripes=False
            )
            tab.tableStyleInfo = style
            ws.add_table(tab)
        
        # Freeze header row
        ws.freeze_panes = 'A2'
        
        logger.info(f"Created sheet '{sheet_name}' with {len(dataframe)} rows")
        return ws
    
    def create_pivot_table_sheet(self, sheet_name, source_data, pivot_config):
        """
        Create a pivot table summary.
        
        Args:
            sheet_name (str): Name of the sheet
            source_data (pd.DataFrame): Source data
            pivot_config (dict): Pivot configuration
        """
        try:
            # Create pivot table using pandas
            index = pivot_config.get('index', [])
            columns = pivot_config.get('columns', None)
            values = pivot_config.get('values', [])
            aggfunc = pivot_config.get('aggfunc', 'sum')
            
            if not index or not values:
                logger.warning(f"Invalid pivot config for '{sheet_name}'")
                return None
            
            pivot_df = pd.pivot_table(
                source_data,
                index=index,
                columns=columns,
                values=values,
                aggfunc=aggfunc,
                fill_value=0
            )
            
            # Flatten multi-level columns if needed
            if isinstance(pivot_df.columns, pd.MultiIndex):
                pivot_df.columns = ['_'.join(map(str, col)).strip() for col in pivot_df.columns.values]
            
            # Reset index to make it a regular dataframe
            pivot_df = pivot_df.reset_index()
            
            # Create sheet with pivot data
            ws = self.create_data_sheet(sheet_name, pivot_df, format_as_table=False)
            
            logger.info(f"Created pivot table '{sheet_name}'")
            return ws
        
        except Exception as e:
            logger.error(f"Error creating pivot table '{sheet_name}': {e}")
            return None
    
    def add_chart(self, sheet_name, chart_config):
        """
        Add a chart to a worksheet.
        
        Args:
            sheet_name (str): Sheet to add chart to
            chart_config (dict): Chart configuration
        """
        try:
            ws = self.workbook[sheet_name]
            
            chart_type = chart_config.get('type', 'bar')
            data_range = chart_config.get('data_range')
            position = chart_config.get('position', 'E2')
            title = chart_config.get('title', 'Chart')
            
            if not data_range:
                logger.warning(f"No data range specified for chart in '{sheet_name}'")
                return
            
            # Determine chart type
            if chart_type == 'bar':
                chart = BarChart()
            elif chart_type == 'pie':
                chart = PieChart()
            elif chart_type == 'line':
                chart = LineChart()
            else:
                logger.warning(f"Unsupported chart type: {chart_type}")
                return
            
            # Set chart title
            chart.title = title
            chart.style = 10
            
            # Add data (simple implementation - can be enhanced)
            # Assuming data_range is like "A1:B10"
            if ':' in data_range:
                data = Reference(ws, range_string=f"{sheet_name}!{data_range}")
                chart.add_data(data, titles_from_data=True)
            
            # Add chart to worksheet
            ws.add_chart(chart, position)
            
            logger.info(f"Added {chart_type} chart to '{sheet_name}' at {position}")
        
        except Exception as e:
            logger.error(f"Error adding chart to '{sheet_name}': {e}")
    
    def create_summary_sheet(self, kpis):
        """
        Create a summary sheet with KPIs.
        
        Args:
            kpis (list): List of KPI configurations
        """
        ws = self.workbook.create_sheet('Summary', 0)
        
        # Title
        ws['A1'] = 'Executive Summary'
        ws['A1'].font = Font(bold=True, size=18, color='FFFFFF')
        ws['A1'].fill = PatternFill(start_color='203764', end_color='203764', fill_type='solid')
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells('A1:D1')
        ws.row_dimensions[1].height = 30
        
        # Report date
        ws['A2'] = 'Report Date:'
        ws['B2'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        ws['A2'].font = Font(bold=True)
        
        # KPIs section
        row = 4
        ws[f'A{row}'] = 'Key Performance Indicators'
        ws[f'A{row}'].font = Font(bold=True, size=14)
        row += 2
        
        # Headers
        ws[f'A{row}'] = 'Metric'
        ws[f'B{row}'] = 'Value'
        ws[f'C{row}'] = 'Target'
        ws[f'D{row}'] = 'Status'
        
        for cell in [ws[f'A{row}'], ws[f'B{row}'], ws[f'C{row}'], ws[f'D{row}']]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
        
        row += 1
        
        # Add KPIs
        for kpi in kpis:
            name = kpi.get('name', 'KPI')
            value = kpi.get('value', 0)
            target = kpi.get('target', 0)
            
            ws[f'A{row}'] = name
            ws[f'B{row}'] = value
            ws[f'C{row}'] = target
            
            # Calculate status
            if isinstance(value, (int, float)) and isinstance(target, (int, float)) and target > 0:
                percentage = (value / target) * 100
                if percentage >= 100:
                    status = '✓ On Target'
                    color = '00B050'
                elif percentage >= 80:
                    status = '⚠ Warning'
                    color = 'FFC000'
                else:
                    status = '✗ Below Target'
                    color = 'FF0000'
                
                ws[f'D{row}'] = status
                ws[f'D{row}'].fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                ws[f'D{row}'].font = Font(color='FFFFFF', bold=True)
            else:
                ws[f'D{row}'] = 'N/A'
            
            ws[f'D{row}'].alignment = Alignment(horizontal='center')
            row += 1
        
        # Auto-adjust columns
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 20
        
        logger.info(f"Created summary sheet with {len(kpis)} KPIs")
    
    def apply_conditional_formatting(self, sheet_name, format_config):
        """
        Apply conditional formatting to a worksheet.
        
        Args:
            sheet_name (str): Sheet name
            format_config (dict): Formatting configuration
        """
        try:
            ws = self.workbook[sheet_name]
            
            format_type = format_config.get('type', 'color_scale')
            cell_range = format_config.get('range')
            
            if not cell_range:
                logger.warning(f"No range specified for conditional formatting in '{sheet_name}'")
                return
            
            if format_type == 'color_scale':
                # Three-color scale: Red -> Yellow -> Green
                rule = ColorScaleRule(
                    start_type='min', start_color='F8696B',
                    mid_type='percentile', mid_value=50, mid_color='FFEB84',
                    end_type='max', end_color='63BE7B'
                )
                ws.conditional_formatting.add(cell_range, rule)
            
            elif format_type == 'data_bars':
                # Data bars
                rule = ColorScaleRule(
                    start_type='min', start_color='63BE7B',
                    end_type='max', end_color='63BE7B'
                )
                ws.conditional_formatting.add(cell_range, rule)
            
            elif format_type == 'highlight_cells':
                # Highlight cells based on value
                operator = format_config.get('operator', 'greaterThan')
                threshold = format_config.get('threshold', 0)
                color = format_config.get('color', 'FFC7CE')
                
                rule = CellIsRule(
                    operator=operator,
                    formula=[threshold],
                    fill=PatternFill(start_color=color, end_color=color, fill_type='solid')
                )
                ws.conditional_formatting.add(cell_range, rule)
            
            logger.info(f"Applied {format_type} formatting to '{sheet_name}' range {cell_range}")
        
        except Exception as e:
            logger.error(f"Error applying conditional formatting: {e}")
    
    def protect_sheet(self, sheet_name, password=None):
        """
        Protect a worksheet with password.
        
        Args:
            sheet_name (str): Sheet name
            password (str): Protection password
        """
        try:
            ws = self.workbook[sheet_name]
            ws.protection.sheet = True
            if password:
                ws.protection.password = password
            logger.info(f"Protected sheet '{sheet_name}'")
        except Exception as e:
            logger.error(f"Error protecting sheet '{sheet_name}': {e}")
    
    def generate_report(self):
        """Generate the complete Excel report."""
        logger.info("=" * 70)
        logger.info("STARTING EXCEL REPORT GENERATION")
        logger.info("=" * 70)
        
        # Load data sources
        self.load_all_data_sources()
        
        if not self.data_sources:
            logger.error("No data sources loaded. Cannot generate report.")
            return None
        
        # Create new workbook
        from openpyxl import Workbook
        self.workbook = Workbook()
        
        # Remove default sheet
        if 'Sheet' in self.workbook.sheetnames:
            self.workbook.remove(self.workbook['Sheet'])
        
        # Create summary sheet with KPIs
        kpis = self.config.get('kpis', [])
        if kpis:
            self.create_summary_sheet(kpis)
        
        # Create data sheets
        for name, df in self.data_sources.items():
            self.create_data_sheet(name, df)
        
        # Create pivot tables
        for pivot in self.config.get('pivot_tables', []):
            sheet_name = pivot.get('sheet_name', 'Pivot')
            source_name = pivot.get('source', list(self.data_sources.keys())[0])
            
            if source_name in self.data_sources:
                self.create_pivot_table_sheet(
                    sheet_name,
                    self.data_sources[source_name],
                    pivot
                )
        
        # Add charts
        for chart in self.config.get('charts', []):
            sheet_name = chart.get('sheet_name')
            if sheet_name and sheet_name in self.workbook.sheetnames:
                self.add_chart(sheet_name, chart)
        
        # Apply conditional formatting
        for fmt in self.config.get('conditional_formatting', []):
            sheet_name = fmt.get('sheet_name')
            if sheet_name and sheet_name in self.workbook.sheetnames:
                self.apply_conditional_formatting(sheet_name, fmt)
        
        # Protect sheets
        password = self.config.get('password')
        if password:
            for sheet_name in self.workbook.sheetnames:
                self.protect_sheet(sheet_name, password)
        
        # Save workbook
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        report_name = self.config.get('report_name', 'automated_report')
        self.output_file = self.output_folder / f"{report_name}_{timestamp}.xlsx"
        
        self.workbook.save(self.output_file)
        logger.info(f"Report saved to: {self.output_file}")
        
        logger.info("=" * 70)
        logger.info("REPORT GENERATION COMPLETE")
        logger.info("=" * 70)
        
        return self.output_file
    
    def send_email_report(self):
        """Send the generated report via email."""
        email_config = self.config.get('email', {})
        
        if not email_config.get('enabled', False):
            logger.info("Email sending is disabled")
            return False
        
        if not self.output_file or not self.output_file.exists():
            logger.error("No report file to send")
            return False
        
        try:
            sender_email = email_config.get('sender_email')
            sender_password = email_config.get('sender_password')
            recipients = email_config.get('recipients', [])
            
            if not sender_email or not sender_password or not recipients:
                logger.error("Email configuration incomplete")
                return False
            
            # Create message
            msg = MIMEMultipart()
            msg['From'] = sender_email
            msg['To'] = ', '.join(recipients)
            msg['Subject'] = email_config.get('subject', 'Automated Excel Report')
            
            # Email body
            body = email_config.get('body', 'Please find the attached report.')
            msg.attach(MIMEText(body, 'plain'))
            
            # Attach file
            with open(self.output_file, 'rb') as attachment:
                part = MIMEBase('application', 'octet-stream')
                part.set_payload(attachment.read())
            
            encoders.encode_base64(part)
            part.add_header(
                'Content-Disposition',
                f'attachment; filename= {self.output_file.name}'
            )
            msg.attach(part)
            
            # Send email
            smtp_server = email_config.get('smtp_server', 'smtp.gmail.com')
            smtp_port = email_config.get('smtp_port', 587)
            
            server = smtplib.SMTP(smtp_server, smtp_port)
            server.starttls()
            server.login(sender_email, sender_password)
            server.send_message(msg)
            server.quit()
            
            logger.info(f"Email sent successfully to {len(recipients)} recipient(s)")
            return True
        
        except Exception as e:
            logger.error(f"Error sending email: {e}")
            return False


def main():
    """Main execution function."""
    
    # Get config file from command line or use default
    if len(sys.argv) > 1:
        config_file = sys.argv[1]
    else:
        config_file = "./config/report_config.json"
        logger.info(f"No config file specified, using default: {config_file}")
    
    try:
        # Create generator
        generator = ExcelReportGenerator(config_file)
        
        # Generate report
        report_file = generator.generate_report()
        
        if report_file:
            print(f"\n✓ Success! Report generated: {report_file}")
            
            # Send email if configured
            if generator.config.get('email', {}).get('enabled', False):
                if generator.send_email_report():
                    print("✓ Email sent successfully")
                else:
                    print("✗ Failed to send email")
        else:
            print("\n✗ Report generation failed")
            sys.exit(1)
    
    except Exception as e:
        logger.error(f"Fatal error: {e}")
        print(f"\n✗ Error: {e}")
        sys.exit(1)


if __name__ == "__main__":
    main()
